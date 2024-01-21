from openpyxl import load_workbook
from difflib import SequenceMatcher
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manager
# 打开 Excel 文件
workbook = load_workbook(filename='creat.xlsx')

# 选择工作表
worksheet = workbook.active

# 读取数据
data = []
for row in worksheet.iter_rows(values_only=True):
    data.append(row[0])

# 关键词
query = '李宁'

# 计算相似度
similarities = [(s, SequenceMatcher(None, s, query).ratio()) for s in data]
similarities.sort(key=lambda x: x[1], reverse=True)

# 取出相似度最高的 6 项
top_similarities = similarities[:6]
print(top_similarities)
# 绘制饼状图
labs = [('李宁跑步鞋男鞋休闲慢跑鞋运动鞋ARST089', 0.16666666666666666), ('李宁天迹丨经典休闲鞋女鞋板鞋运动鞋AGCT376', 0.15384615384615385), ('李宁江月丨板鞋女鞋减震回弹经典休闲鞋AGCT294', 0.14814814814814814), ('李宁漫云丨跑步鞋女鞋休闲慢跑鞋运动鞋ARST068', 0.14814814814814814), ('李宁CF溯誉丨板鞋男鞋经典休闲鞋运动鞋AGCT329', 0.14285714285714285), ('李宁元宝丨经典休闲鞋女鞋板鞋低帮运动鞋AGCT340', 0.14285714285714285)]
labels = [s[0] for s in labs]
sizes = [s[1] for s in top_similarities]
plt.pie(sizes, labels=labels, autopct='%1.1f%%')
plt.title('Top 6 相似度')
plt.show()

# 绘制柱状图
plt.bar(labels, sizes)
plt.title('Top 6 Similarities')
plt.xlabel(top_similarities)
plt.ylabel('Similarity')
plt.show()
