import pandas as pd
import matplotlib.pyplot as plt
import time
import openpyxl
import requests
from bs4 import BeautifulSoup
import pandas as pd
from fake_useragent import UserAgent


ua = UserAgent()


workbook = openpyxl.Workbook()

# 获取活动工作表
worksheet = workbook.active
# 写入列标题

headers = {
'User-Agent': ua.random,
    "Cookie":'ASP.NET_SessionId=rrowqwvad0o3oc45y3oyfi55; Hm_lvt_f48cedd6a69101030e93d4ef60f48fd0=1715782911,1715870021; __51cke__=; __gads=ID=d081f14b139332e2:T=1715782910:RT=1715870019:S=ALNI_MbGeAQIUXbFLGIRkAAA61vGUaPecg; __gpi=UID=00000e1e55416bc2:T=1715782910:RT=1715870019:S=ALNI_MZpwhxIqv-X7OHQZohXuE8ZGHS2YQ; __eoi=ID=c66282bd0d775649:T=1715782910:RT=1715870019:S=AA-AfjYgYPhaX0jXIHLtPvejjl4S; __tins__21287555=%7B%22sid%22%3A%201715870021302%2C%20%22vd%22%3A%205%2C%20%22expires%22%3A%201715872046443%7D; __51laig__=5; Hm_lpvt_f48cedd6a69101030e93d4ef60f48fd0=1715870246'
}

def get_data():
    df = pd.read_excel('保定天气link.xlsx')
    column_name = 'ln'
    links = df[column_name]

    worksheet.cell(row=1, column=1, value='日期')
    worksheet.cell(row=1, column=2, value='天气状况')
    worksheet.cell(row=1, column=3, value='最低气温/最高气温')
    worksheet.cell(row=1, column=4, value='风力风向(夜间/白天)')

    # 循环处理每个链接
    for link in links:
        time.sleep(3)
        response = requests.get(url=link, headers=headers).text
        soup = BeautifulSoup(response, 'html.parser')
        td_elements = soup.find_all('td')
        td_values = [td.get_text(strip=True).replace('\r\n', '').strip() for td in td_elements]
        data = []
        for i in range(0, len(td_values), 8):
            # 从第五个元素开始提取数据
            row = [val if val else '-' for val in td_values[i+4:i+8]]
            data.append(row)

        # 获取当前工作表的最后一行
        last_row = worksheet.max_row

        # 写入数据
        row = last_row + 1
        for data_row in data:
            col = 1
            for item in data_row:
                worksheet.cell(row=row, column=col, value=item)
                col += 1
            row += 1

        # 保存工作簿
        workbook.save('output.xlsx')
        print(f'Successfully processed: {link}')

def plt_pie():
    df = pd.read_excel('output.xlsx')

    # 提取日期、最低气温和最高气温
    dates = df['日期'].tolist()
    min_temps = df['最低气温'].tolist()
    max_temps = df['最高气温'].tolist()

    # 创建一个包含日期和温度的字典
    date_temp_dict = {dates[i]: {'min_temp': min_temps[i], 'max_temp': max_temps[i]} for i in range(len(dates))}

    # 计算最高气温和最低气温的频率
    max_temp_freq = pd.Series([date_temp_dict[date]['max_temp'] for date in dates]).value_counts()
    min_temp_freq = pd.Series([date_temp_dict[date]['min_temp'] for date in dates]).value_counts()

    # 获取出现次数最多的10个最高气温和最低气温
    top_10_max_temps = max_temp_freq.head(10)
    top_10_min_temps = min_temp_freq.head(10)

    # 绘制最高气温的饼状图
    plt.figure(figsize=(10, 8))
    plt.pie(top_10_max_temps, labels=top_10_max_temps.index, autopct='%1.1f%%')
    plt.title('Top 10 最高气温')
    plt.show()

    # 绘制最低气温的饼状图
    plt.figure(figsize=(10, 8))
    plt.pie(top_10_min_temps, labels=top_10_min_temps.index, autopct='%1.1f%%')
    plt.title('Top 10 最低气温')
    plt.show()

def plt_bar():
    df = pd.read_excel('output.xlsx')

    # 提取日期、最低气温和最高气温
    dates = df['日期'].tolist()
    min_temps = df['最低气温'].tolist()
    max_temps = df['最高气温'].tolist()

    # 创建一个包含日期和温度的字典
    date_temp_dict = {dates[i]: {'min_temp': min_temps[i], 'max_temp': max_temps[i]} for i in range(len(dates))}

    # 计算最高气温和最低气温的出现次数
    max_temp_counts = pd.Series([date_temp_dict[date]['max_temp'] for date in dates]).value_counts()
    min_temp_counts = pd.Series([date_temp_dict[date]['min_temp'] for date in dates]).value_counts()

    # 获取出现次数最多的10个最高气温和最低气温
    top_10_max_temps = max_temp_counts.head(10)
    top_10_min_temps = min_temp_counts.head(10)

    # 绘制最高气温的柱状图
    plt.figure(figsize=(10, 6))
    plt.bar(range(len(top_10_max_temps)), top_10_max_temps.values, color='skyblue')
    plt.xticks(ticks=range(len(top_10_max_temps)), labels=top_10_max_temps.index, rotation=45)
    plt.xlabel('最高气温 (°C)')
    plt.ylabel('出现次数')
    plt.title('Top 10 Highest Temperatures')
    plt.show()

    # 绘制最低气温的柱状图
    plt.figure(figsize=(10, 6))
    plt.bar(range(len(top_10_min_temps)), top_10_min_temps.values, color='lightcoral')
    plt.xticks(ticks=range(len(top_10_min_temps)), labels=top_10_min_temps.index, rotation=45)
    plt.xlabel('最低气温 (°C)')
    plt.ylabel('出现次数')
    plt.title('Top 10 Lowest Temperatures')
    plt.show()

def descr():
    # 创建一个 DataFrame 对象
    df = pd.read_excel('output.xlsx')

    # 将温度列转换为数字类型（去掉℃符号）
    df['最低气温'] = df['最低气温'].str.replace('℃', '').astype(int)
    df['最高气温'] = df['最高气温'].str.replace('℃', '').astype(int)

    # 获取最低气温的最小值和最高气温的最大值
    min_low_temp = df['最低气温'].min()
    max_high_temp = df['最高气温'].max()

    # 打印结果
    print(f"最低气温的最小值为: {min_low_temp}℃")
    print(f"最高气温的最大值为: {max_high_temp}℃")

    # 数据用于条形图
    values = [min_low_temp, max_high_temp]
    labels = ['最低气温的最小值', '最高气温的最大值']

    # 创建条形图
    plt.figure(figsize=(8, 6))
    plt.bar(labels, values, color=['blue', 'red'])

    # 添加标题和标签
    plt.title('最低气温的最小值和最高气温的最大值')
    plt.ylabel('温度 (℃)')

    # 显示图表
    plt.show()
if __name__ == '__main__':
    descr()